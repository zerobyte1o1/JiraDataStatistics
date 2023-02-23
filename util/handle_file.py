import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


class Handlefile:
    def read_execl(self, filename, sheetname):
        wb = openpyxl.load_workbook(filename)
        # 指定要读取内容的sheet
        sheetobj = wb[sheetname]
        # 读取所有的内容
        # 获取最大行
        rows = sheetobj.max_row
        # 获取最大列
        columns = sheetobj.max_column
        list = []
        # 使用for循环，
        for row in range(2, rows + 1):
            # 定义一个列表，用来存放我们获取的每一个数据
            new_list = []
            # 列的处理
            for column in range(1, columns + 1):
                # 取值
                new_list.append(sheetobj.cell(row, column).value)
            list.append(new_list)
        return list

    def write_excel(self, header: list, data: list):
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.row_dimensions[1].height = 20
        for i in range(len(header)):
            sheet.cell(1, i + 1).value = header[i]
            sheet.cell(1, i + 1).font = Font(bold=True, size=15)

        for j in range(len(data)):
            for k in range(len(data[j])):
                sheet.cell(j + 2, k + 1).value = data[j][k]
                sheet.cell(j + 2, k + 1).alignment=Alignment(wrapText=True)
        return wb,sheet

#
# if __name__ == '__main__':
#     wb = write_excel(["A", "b"], [[1, 2], [3, 4]])
#     print(wb)
